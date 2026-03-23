from __future__ import annotations

# ── License check (runs FIRST, before any GUI opens) ────────────────────────
from license_client import check_license, get_college_name
check_license()
# ────────────────────────────────────────────────────────────────────────────

import os
import sys
import threading
from typing import Dict, List

import customtkinter as ctk
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkinter import filedialog, messagebox, ttk

from pdf_parser import parse_cr_pdf, detect_input_type

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

BG_MAIN = "#0F172A"
BG_PANEL = "#111827"
BG_CARD = "#1F2937"
BG_CARD_2 = "#243041"
BG_LOG = "#0B1220"
BORDER = "#334155"
TEXT_MAIN = "#F8FAFC"
TEXT_DIM = "#94A3B8"
PRIMARY = "#2563EB"
PRIMARY_HOVER = "#1D4ED8"
SUCCESS = "#16A34A"
SUCCESS_HOVER = "#15803D"
DANGER = "#DC2626"
WARNING = "#D97706"
INFO = "#0891B2"
INFO_HOVER = "#0E7490"
MUTED_BTN = "#374151"
MUTED_HOVER = "#4B5563"
HIGHLIGHT = "#38BDF8"


def force_maximize_window(root):
    try:
        root.update_idletasks()
        root.state("zoomed")
        return
    except Exception:
        pass

    if sys.platform.startswith("win"):
        try:
            import ctypes
            user32 = ctypes.windll.user32
            SW_MAXIMIZE = 3
            root.update_idletasks()
            hwnd = user32.GetParent(root.winfo_id())
            user32.ShowWindow(hwnd, SW_MAXIMIZE)
            return
        except Exception:
            pass

    root.update_idletasks()
    root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")


class StatCard(ctk.CTkFrame):
    def __init__(self, master, title: str, icon: str, value_color: str):
        super().__init__(master, fg_color=BG_CARD, corner_radius=20, border_width=1, border_color=BORDER)
        self.grid_columnconfigure(1, weight=1)

        self.icon_lbl = ctk.CTkLabel(
            self,
            text=icon,
            width=54,
            height=54,
            corner_radius=14,
            fg_color=BG_CARD_2,
            text_color=HIGHLIGHT,
            font=ctk.CTkFont(size=24, weight="bold"),
        )
        self.icon_lbl.grid(row=0, column=0, rowspan=2, padx=(14, 10), pady=14)

        self.title_lbl = ctk.CTkLabel(
            self,
            text=title,
            text_color=TEXT_DIM,
            anchor="w",
            font=ctk.CTkFont(size=12, weight="bold"),
        )
        self.title_lbl.grid(row=0, column=1, sticky="sw", padx=(0, 14), pady=(14, 2))

        self.value_lbl = ctk.CTkLabel(
            self,
            text="--",
            text_color=value_color,
            anchor="w",
            font=ctk.CTkFont(size=28, weight="bold"),
        )
        self.value_lbl.grid(row=1, column=1, sticky="nw", padx=(0, 14), pady=(0, 14))

    def set_value(self, value: str):
        self.value_lbl.configure(text=value)


class ResultAnalyzerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("College Result Analyzer")
        self.configure(fg_color=BG_MAIN)
        self.minsize(1100, 700)
        self.after(80, lambda: force_maximize_window(self))

        self.files: List[str] = []
        self.df = pd.DataFrame()
        self.meta: Dict = {}
        self.analysis: Dict = {}
        self.subject_options: List[str] = []
        self.selected_subjects: List[str] = []

        self._build_layout()

    # ---------- Layout ----------
    def _build_layout(self):
        self.sidebar = ctk.CTkFrame(self, width=250, corner_radius=0, fg_color=BG_PANEL)
        self.sidebar.pack(side="left", fill="y")

        brand = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        brand.pack(fill="x", padx=18, pady=(18, 12))
        ctk.CTkLabel(
            brand, text="⚡ Result Analyzer", text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=26, weight="bold")
        ).pack(anchor="w")
        ctk.CTkLabel(
            brand, text="Professional Result Dashboard", text_color=TEXT_DIM,
            font=ctk.CTkFont(size=12)
        ).pack(anchor="w", pady=(2, 0))
        ctk.CTkFrame(self.sidebar, height=2, fg_color=BORDER).pack(fill="x", padx=18, pady=(0, 16))

        def side_btn(text, cmd, state="normal", color=PRIMARY, hover=PRIMARY_HOVER):
            return ctk.CTkButton(
                self.sidebar,
                text=text,
                command=cmd,
                state=state,
                fg_color=color,
                hover_color=hover,
                corner_radius=14,
                height=44,
                text_color="white",
                font=ctk.CTkFont(size=14, weight="bold"),
            )

        self.btn_select = side_btn("📄 Select Files", self.select_files)
        self.btn_select.pack(fill="x", padx=18, pady=6)

        self.btn_process = side_btn("⚙ Process & Analyze", self.process_and_analyze, state="disabled")
        self.btn_process.pack(fill="x", padx=18, pady=6)

        self.btn_graph = side_btn("📊 Subject Graphs", self.open_graph_page, state="disabled", color=INFO, hover=INFO_HOVER)
        self.btn_graph.pack(fill="x", padx=18, pady=6)

        self.btn_top10 = side_btn("🏆 View Top 10", self.open_top10_page, state="disabled", color=INFO, hover=INFO_HOVER)
        self.btn_top10.pack(fill="x", padx=18, pady=6)

        self.btn_students = side_btn("📋 View Students", self.open_students_page, state="disabled", color=SUCCESS, hover=SUCCESS_HOVER)
        self.btn_students.pack(fill="x", padx=18, pady=6)

        self.btn_export_excel = side_btn("⬇ Export Excel", self.export_excel, state="disabled", color=PRIMARY, hover=PRIMARY_HOVER)
        self.btn_export_excel.pack(fill="x", padx=18, pady=6)

        self.btn_clear = ctk.CTkButton(
            self.sidebar,
            text="🧹 Clear",
            command=self.clear_all,
            fg_color=MUTED_BTN,
            hover_color=MUTED_HOVER,
            corner_radius=14,
            height=42,
            text_color="white",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        self.btn_clear.pack(fill="x", padx=18, pady=(14, 10))

        foot = ctk.CTkFrame(self.sidebar, fg_color=BG_CARD, corner_radius=16, border_width=1, border_color=BORDER)
        foot.pack(side="bottom", fill="x", padx=18, pady=18)
        ctk.CTkLabel(foot, text="Workflow", text_color=TEXT_DIM, font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=14, pady=(12, 4))
        ctk.CTkLabel(
            foot,
            text="1. Select PDFs\n2. Process Analysis\n3. Open graphs / pages\n4. Export Excel",
            justify="left",
            text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=13),
        ).pack(anchor="w", padx=14, pady=(0, 12))

        self.main = ctk.CTkFrame(self, fg_color=BG_MAIN)
        self.main.pack(side="left", fill="both", expand=True, padx=14, pady=14)
        self.page_container = ctk.CTkFrame(self.main, fg_color=BG_MAIN)
        self.page_container.pack(fill="both", expand=True)

        self.pages = {}
        self._create_home_page()
        self._create_graph_page()
        self._create_top10_page()
        self._create_students_page()
        self._create_student_details_page()
        self.show_page("home")

        try:
            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview", background="#0B1220", foreground="white", fieldbackground="#0B1220", rowheight=28)
            style.configure("Treeview.Heading", background="#1F2937", foreground="white")
            style.map("Treeview", background=[("selected", "#1D4ED8")], foreground=[("selected", "white")])
        except Exception:
            pass

    def _page_header(self, parent, title, back_to="home"):
        header = ctk.CTkFrame(parent, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        header.pack(fill="x", padx=12, pady=(12, 10))
        ctk.CTkButton(
            header,
            text="⬅ Back",
            fg_color=MUTED_BTN,
            hover_color=MUTED_HOVER,
            width=90,
            corner_radius=12,
            command=lambda: self.show_page(back_to),
        ).pack(side="left", padx=12, pady=12)
        ctk.CTkLabel(
            header, text=title, text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(side="left", padx=12)

    def show_page(self, name):
        for p in self.pages.values():
            p.pack_forget()
        self.pages[name].pack(fill="both", expand=True)

    # ---------- Home ----------
    def _create_home_page(self):
        home = ctk.CTkFrame(self.page_container, fg_color=BG_MAIN)
        self.pages["home"] = home

        hero = ctk.CTkFrame(home, fg_color=BG_PANEL, corner_radius=20, border_width=1, border_color=BORDER)
        hero.pack(fill="x", padx=12, pady=(12, 12))
        hero.grid_columnconfigure(1, weight=1)

        info = ctk.CTkFrame(hero, fg_color=BG_CARD, corner_radius=18, border_width=1, border_color=BORDER)
        info.grid(row=0, column=0, sticky="nsw", padx=16, pady=16)
        ctk.CTkLabel(info, text="📂 Quick Start", text_color=TEXT_MAIN, font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", padx=16, pady=(16, 6))
        ctk.CTkLabel(
            info,
            text="• Click Select Files\n• Click Process & Analyze\n• Open Subject Graphs / Students / Top 10\n• Export Excel report",
            justify="left",
            text_color=TEXT_DIM,
            font=ctk.CTkFont(size=12),
        ).pack(anchor="w", padx=16, pady=(0, 16))

        cards_wrap = ctk.CTkFrame(hero, fg_color="transparent")
        cards_wrap.grid(row=0, column=1, sticky="nsew", padx=(0, 16), pady=16)
        for i in range(4):
            cards_wrap.grid_columnconfigure(i, weight=1)

        self.card_students = StatCard(cards_wrap, "Students", "👥", TEXT_MAIN)
        self.card_students.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
        self.card_pass = StatCard(cards_wrap, "Pass", "✓", SUCCESS)
        self.card_pass.grid(row=0, column=1, sticky="ew", padx=6, pady=6)
        self.card_fail = StatCard(cards_wrap, "Fail", "✕", DANGER)
        self.card_fail.grid(row=0, column=2, sticky="ew", padx=6, pady=6)
        self.card_avg = StatCard(cards_wrap, "Average %", "◫", PRIMARY)
        self.card_avg.grid(row=0, column=3, sticky="ew", padx=6, pady=6)

        log_card = ctk.CTkFrame(home, fg_color=BG_PANEL, corner_radius=20, border_width=1, border_color=BORDER)
        log_card.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        ctk.CTkLabel(log_card, text="Processing Log", text_color=TEXT_MAIN, font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w", padx=16, pady=(14, 8))
        self.logbox = ctk.CTkTextbox(log_card, height=330, fg_color=BG_LOG, text_color=HIGHLIGHT, corner_radius=14, border_width=1, border_color=BORDER)
        self.logbox.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        self.log("Ready. 1) Select PDF(s)  2) Process")

        self.status = ctk.CTkLabel(home, text="Status: Idle", text_color=HIGHLIGHT, font=ctk.CTkFont(size=13, weight="bold"))
        self.status.pack(anchor="w", padx=16, pady=(0, 12))

        self.lbl_students = self.card_students.value_lbl
        self.lbl_pass = self.card_pass.value_lbl
        self.lbl_fail = self.card_fail.value_lbl
        self.lbl_avg = self.card_avg.value_lbl

    # ---------- Graphs ----------
    def _create_graph_page(self):
        page = ctk.CTkFrame(self.page_container, fg_color=BG_MAIN)
        self.pages["graphs"] = page
        self._page_header(page, "📊 Subject Wise Graphs", back_to="home")

        top = ctk.CTkFrame(page, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        top.pack(fill="x", padx=12, pady=(0, 12))
        ctk.CTkLabel(top, text="Choose Subject:", text_color=TEXT_MAIN, font=ctk.CTkFont(size=14, weight="bold")).pack(side="left", padx=(16, 10), pady=14)

        self.graph_subject_var = ctk.StringVar(value="All Subjects")
        self.graph_subject_menu = ctk.CTkOptionMenu(
            top,
            values=["All Subjects"],
            variable=self.graph_subject_var,
            command=lambda _v: self.render_subject_graph(),
            fg_color=PRIMARY,
            button_color=PRIMARY,
            button_hover_color=PRIMARY_HOVER,
            width=320,
        )
        self.graph_subject_menu.pack(side="left", padx=(0, 12), pady=12)

        self.graph_hint = ctk.CTkLabel(top, text="Process PDFs first, then choose a subject.", text_color=TEXT_DIM)
        self.graph_hint.pack(side="left", padx=10)

        self.graph_card = ctk.CTkFrame(page, fg_color=BG_PANEL, corner_radius=20, border_width=1, border_color=BORDER)
        self.graph_card.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    def open_graph_page(self):
        self.show_page("graphs")
        self.render_subject_graph()

    def render_subject_graph(self):
        for w in self.graph_card.winfo_children():
            w.destroy()

        if self.df.empty:
            ctk.CTkLabel(self.graph_card, text="No analysis data available.", text_color=TEXT_DIM, font=ctk.CTkFont(size=18, weight="bold")).pack(expand=True)
            return

        choice = self.graph_subject_var.get()

        fig = Figure(figsize=(8, 4.6), dpi=100)
        fig.patch.set_facecolor(BG_PANEL)
        ax = fig.add_subplot(111)
        ax.set_facecolor(BG_PANEL)
        ax.tick_params(colors=TEXT_DIM)
        for spine in ax.spines.values():
            spine.set_color(BORDER)

        if choice == "All Subjects":
            names = []
            values = []
            for info in self._subject_info_list_export(self._subject_groups()):
                vals = pd.to_numeric(self.df.get(info["tot_col"], pd.Series(dtype=float)), errors="coerce")
                avg = vals.mean()
                if pd.notna(avg):
                    names.append(info["display_name"][:28])
                    values.append(float(avg))

            if not names:
                ax.text(0.5, 0.5, "No subject totals found", color=TEXT_DIM, ha="center", va="center", fontsize=14)
                ax.set_xticks([])
                ax.set_yticks([])
            else:
                ypos = list(range(len(names)))
                bars = ax.barh(ypos, values, color=PRIMARY, edgecolor=HIGHLIGHT, linewidth=1.0)
                ax.set_yticks(ypos)
                ax.set_yticklabels(names, color=TEXT_DIM)
                ax.invert_yaxis()
                ax.set_xlabel("Average Marks", color=TEXT_DIM)
                ax.set_title("All Subjects - Average Total Marks", color=TEXT_MAIN, fontsize=14, fontweight="bold")
                ax.grid(axis="x", linestyle="--", alpha=0.25, color=TEXT_DIM)
                for bar, val in zip(bars, values):
                    ax.text(val + 0.3, bar.get_y() + bar.get_height() / 2, f"{val:.1f}", va="center", color=TEXT_MAIN, fontsize=9)
        else:
            info = None
            for x in self._subject_info_list_export(self._subject_groups()):
                if x["display_name"] == choice:
                    info = x
                    break

            if info is None:
                ax.text(0.5, 0.5, "Subject not found", color=TEXT_DIM, ha="center", va="center", fontsize=14)
                ax.set_xticks([])
                ax.set_yticks([])
            else:
                status_series = self._subject_status_dataframe_export(self.df, self._subject_groups())[info["status_col"]].astype(str).str.upper()
                counts = [
                    int((status_series == "PASS").sum()),
                    int((status_series == "FAIL").sum()),
                    int((status_series == "ABSENT").sum()),
                ]
                labels = ["Pass", "Fail", "Absent"]
                bars = ax.bar(labels, counts, color=[SUCCESS, DANGER, WARNING], edgecolor=HIGHLIGHT, linewidth=1.0)
                ax.set_title(f"{choice} - Pass / Fail / Absent", color=TEXT_MAIN, fontsize=14, fontweight="bold")
                ax.set_ylabel("Students", color=TEXT_DIM)
                ax.grid(axis="y", linestyle="--", alpha=0.25, color=TEXT_DIM)
                for bar, val in zip(bars, counts):
                    ax.text(bar.get_x() + bar.get_width() / 2, val + 0.2, str(val), ha="center", color=TEXT_MAIN, fontsize=10, fontweight="bold")

        canvas = FigureCanvasTkAgg(fig, master=self.graph_card)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=14, pady=14)

    # ---------- Top 10 ----------
    def _create_top10_page(self):
        page = ctk.CTkFrame(self.page_container, fg_color=BG_MAIN)
        self.pages["top10"] = page
        self._page_header(page, "🏆 Top 10 Students", back_to="home")
        self.top10_scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        self.top10_scroll.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    def open_top10_page(self):
        for w in self.top10_scroll.winfo_children():
            w.destroy()

        top10 = self.analysis.get("top10", [])
        if not top10:
            ctk.CTkLabel(self.top10_scroll, text="No top 10 data available.", text_color=TEXT_MAIN).pack(padx=12, pady=12)
            self.show_page("top10")
            return

        for i, r in enumerate(top10, start=1):
            name = (r.get("name") or "").strip() or f"(Name missing) - {r.get('enrollment_no', 'NA')}"
            enroll = r.get("enrollment_no", "NA")
            marks = r.get("total_marks_550", "NA")
            pct = r.get("percentage", "NA")
            sgpa = r.get("sgpa", "NA")
            res = (r.get("result") or "UNKNOWN").upper()
            res_color = SUCCESS if res == "PASS" else DANGER if res == "FAIL" else PRIMARY

            card = ctk.CTkFrame(self.top10_scroll, fg_color=BG_CARD, corner_radius=18, border_width=1, border_color=BORDER)
            card.pack(fill="x", padx=12, pady=10)
            top_row = ctk.CTkFrame(card, fg_color=BG_CARD)
            top_row.pack(fill="x", padx=14, pady=(12, 6))
            ctk.CTkLabel(top_row, text=f"{i:02d}", text_color="white", fg_color=PRIMARY, corner_radius=10, width=44, height=28).pack(side="left")
            ctk.CTkLabel(top_row, text=name, text_color=TEXT_MAIN, font=ctk.CTkFont(size=16, weight="bold")).pack(side="left", padx=12)
            ctk.CTkLabel(top_row, text=res, text_color="white", fg_color=res_color, corner_radius=10, width=70, height=28).pack(side="right")

            details = ctk.CTkFrame(card, fg_color=BG_CARD_2, corner_radius=12)
            details.pack(fill="x", padx=14, pady=(0, 12))
            ctk.CTkLabel(details, text=f"Enroll: {enroll}", text_color="white").pack(anchor="w", padx=12, pady=(10, 4))
            ctk.CTkLabel(details, text=f"Marks: {marks}/550   |   %: {pct}   |   SGPA: {sgpa}", text_color=TEXT_DIM).pack(anchor="w", padx=12, pady=(0, 10))

        self.show_page("top10")

    # ---------- Students ----------
    def _create_students_page(self):
        page = ctk.CTkFrame(self.page_container, fg_color=BG_MAIN)
        self.pages["students"] = page
        self._page_header(page, "📋 Students List (Double-click for Details)", back_to="home")
        container = ctk.CTkFrame(page, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        container.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        bar = ctk.CTkFrame(container, fg_color=BG_CARD, corner_radius=12)
        bar.pack(fill="x", padx=12, pady=12)
        ctk.CTkLabel(bar, text="Search:", text_color="white").pack(side="left", padx=(12, 6), pady=10)
        self.search_var = ctk.StringVar()
        self.search_ent = ctk.CTkEntry(bar, width=250, textvariable=self.search_var, placeholder_text="Enrollment / Name")
        self.search_ent.pack(side="left", padx=(0, 12), pady=10)
        ctk.CTkLabel(bar, text="Result:", text_color="white").pack(side="left", padx=(6, 6), pady=10)
        self.result_var = ctk.StringVar(value="ALL")
        self.result_opt = ctk.CTkOptionMenu(bar, values=["ALL", "PASS", "FAIL", "ABSENT", "UNKNOWN"], variable=self.result_var)
        self.result_opt.pack(side="left", padx=(0, 12), pady=10)
        self.btn_apply = ctk.CTkButton(bar, text="Apply", fg_color=PRIMARY, hover_color=PRIMARY_HOVER, command=self.refresh_students_table)
        self.btn_apply.pack(side="right", padx=(6, 12), pady=10)
        table_frame = ctk.CTkFrame(container, fg_color=BG_LOG, corner_radius=16, border_width=1, border_color=BORDER)
        table_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        cols = ("enrollment_no", "name", "result", "total", "percentage", "sgpa")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=20)
        for col, txt, width in [
            ("enrollment_no", "Seat / Enrollment", 160),
            ("name", "Name", 420),
            ("result", "Result", 90),
            ("total", "Total(550)", 90),
            ("percentage", "%", 90),
            ("sgpa", "SGPA", 90),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=width)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=10)
        vsb.grid(row=0, column=1, sticky="ns", padx=(0, 10), pady=10)
        hsb.grid(row=1, column=0, sticky="ew", padx=(10, 0), pady=(0, 10))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        self.tree.bind("<Double-1>", self._open_selected_student_details)

    def open_students_page(self):
        if self.df.empty:
            messagebox.showwarning("Students", "No data loaded.")
            return
        self.refresh_students_table()
        self.show_page("students")

    def refresh_students_table(self):
        if self.df.empty:
            return
        self.tree.delete(*self.tree.get_children())
        df = self.df.copy()

        s = self.search_var.get().strip().upper()
        if s:
            df["__name_u"] = df["name"].astype(str).str.upper()
            df["__en_u"] = df["enrollment_no"].astype(str).str.upper()
            df = df[df["__name_u"].str.contains(s, na=False) | df["__en_u"].str.contains(s, na=False)]
            df = df.drop(columns=["__name_u", "__en_u"], errors="ignore")

        r = self.result_var.get()
        if r != "ALL":
            df = df[~df["result"].isin(["PASS", "FAIL", "ABSENT"])] if r == "UNKNOWN" else df[df["result"] == r]

        if "enrollment_no" in df.columns:
            df = df.sort_values("enrollment_no", ascending=True)

        for _, row in df.iterrows():
            self.tree.insert("", "end", values=(
                row.get("enrollment_no", "NA"),
                row.get("name", "NA"),
                row.get("result", "NA"),
                row.get("total_marks_550", "NA"),
                row.get("percentage", "NA"),
                row.get("sgpa", "NA"),
            ))

    # ---------- Details ----------
    def _create_student_details_page(self):
        page = ctk.CTkFrame(self.page_container, fg_color=BG_MAIN)
        self.pages["details"] = page
        self._page_header(page, "👤 Student Details", back_to="students")

        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(1, weight=1)

        self.details_summary = ctk.CTkFrame(page, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        self.details_summary.pack(fill="x", padx=12, pady=(0, 12))

        self.details_stats_wrap = ctk.CTkFrame(self.details_summary, fg_color="transparent")
        self.details_stats_wrap.pack(fill="x", padx=14, pady=14)
        for i in range(5):
            self.details_stats_wrap.grid_columnconfigure(i, weight=1)

        self.detail_name = ctk.CTkLabel(
            self.details_summary,
            text="Select a student from the Students page",
            text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.detail_name.pack(anchor="w", padx=16, pady=(0, 14))

        self.detail_stat_result = StatCard(self.details_stats_wrap, "Result", "✓", PRIMARY)
        self.detail_stat_result.grid(row=0, column=0, sticky="ew", padx=6, pady=4)
        self.detail_stat_total = StatCard(self.details_stats_wrap, "Total", "◫", TEXT_MAIN)
        self.detail_stat_total.grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        self.detail_stat_pct = StatCard(self.details_stats_wrap, "Percentage", "%", PRIMARY)
        self.detail_stat_pct.grid(row=0, column=2, sticky="ew", padx=6, pady=4)
        self.detail_stat_sgpa = StatCard(self.details_stats_wrap, "SGPA", "◎", INFO)
        self.detail_stat_sgpa.grid(row=0, column=3, sticky="ew", padx=6, pady=4)
        self.detail_stat_cgpa = StatCard(self.details_stats_wrap, "CGPA", "◆", WARNING)
        self.detail_stat_cgpa.grid(row=0, column=4, sticky="ew", padx=6, pady=4)

        self.details_card = ctk.CTkFrame(page, fg_color=BG_PANEL, corner_radius=18, border_width=1, border_color=BORDER)
        self.details_card.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        ctk.CTkLabel(
            self.details_card,
            text="Subject Breakdown",
            text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(anchor="w", padx=16, pady=(14, 8))

        self.details_scroll = ctk.CTkScrollableFrame(self.details_card, fg_color=BG_LOG, corner_radius=14, border_width=1, border_color=BORDER)
        self.details_scroll.pack(fill="both", expand=True, padx=16, pady=(0, 16))

    def _detail_card(self, parent, title: str, the_v: str, int_v: str, pra_v: str, total_v: str, grade_v: str, status_v: str):
        row = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=16, border_width=1, border_color=BORDER)
        row.pack(fill="x", padx=8, pady=8)

        top = ctk.CTkFrame(row, fg_color="transparent")
        top.pack(fill="x", padx=14, pady=(12, 8))

        status_upper = str(status_v).upper()
        badge_color = SUCCESS if status_upper == "PASS" else DANGER if status_upper == "FAIL" else WARNING if status_upper == "ABSENT" else PRIMARY

        ctk.CTkLabel(
            top,
            text=title,
            text_color=TEXT_MAIN,
            font=ctk.CTkFont(size=15, weight="bold")
        ).pack(side="left")

        ctk.CTkLabel(
            top,
            text=status_v or "NA",
            text_color="white",
            fg_color=badge_color,
            corner_radius=10,
            width=80,
            height=28,
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="right")

        grid = ctk.CTkFrame(row, fg_color=BG_CARD_2, corner_radius=12)
        grid.pack(fill="x", padx=14, pady=(0, 12))
        for i in range(5):
            grid.grid_columnconfigure(i, weight=1)

        items = [
            ("THE", the_v),
            ("INT", int_v),
            ("PRA", pra_v if str(pra_v).strip() else "-"),
            ("TOTAL", total_v),
            ("GRADE", grade_v if str(grade_v).strip() else "-"),
        ]
        for col, (label, value) in enumerate(items):
            box = ctk.CTkFrame(grid, fg_color=BG_CARD, corner_radius=10)
            box.grid(row=0, column=col, sticky="ew", padx=6, pady=8)
            ctk.CTkLabel(box, text=label, text_color=TEXT_DIM, font=ctk.CTkFont(size=11, weight="bold")).pack(anchor="w", padx=10, pady=(8, 2))
            ctk.CTkLabel(box, text=str(value), text_color=TEXT_MAIN, font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10, pady=(0, 8))

    def _open_selected_student_details(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        enroll = str(vals[0]).strip()
        row = self.df[self.df["enrollment_no"].astype(str).str.strip() == enroll]
        if row.empty:
            return
        r = row.iloc[0]

        for w in self.details_scroll.winfo_children():
            w.destroy()

        student_name = str(r.get("name", "NA"))
        self.detail_name.configure(text=f"{student_name}  •  Enrollment: {r.get('enrollment_no','NA')}")

        result_value = str(r.get("result", "NA"))
        self.detail_stat_result.set_value(result_value)
        try:
            self.detail_stat_result.value_lbl.configure(text_color=SUCCESS if result_value.upper() == "PASS" else DANGER if result_value.upper() == "FAIL" else WARNING)
        except Exception:
            pass
        self.detail_stat_total.set_value(str(r.get("total_marks_550", "NA")))
        self.detail_stat_pct.set_value(str(r.get("percentage", "NA")))
        self.detail_stat_sgpa.set_value(str(r.get("sgpa", "NA")))
        self.detail_stat_cgpa.set_value(str(r.get("cgpa", "NA")))

        one = self._subject_status_dataframe_export(row, self._subject_groups())
        rr = one.iloc[0] if not one.empty else r

        for info in self._subject_info_list_export(self._subject_groups()):
            self._detail_card(
                self.details_scroll,
                info["display_name"],
                rr.get(info["the_col"], ""),
                rr.get(info["int_col"], ""),
                rr.get(info["pra_col"], "") if info["pra_col"] in one.columns else "",
                rr.get(info["tot_col"], ""),
                rr.get(info["grade_col"], ""),
                rr.get(info["status_col"], ""),
            )

        self.show_page("details")

    # ---------- Actions ----------

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select Result Files (PDF)",
            filetypes=[("PDF Files", "*.pdf"), ("All Supported", "*.pdf")]
        )
        if not files:
            return
        self.files = list(files)
        self.btn_process.configure(state="normal")
        self.set_status("Files selected")
        self.log(f"\nSelected {len(self.files)} file(s):")
        for f in self.files:
            self.log(f" - {os.path.basename(f)}  [{detect_input_type(f)}]")

    def process_and_analyze(self):
        if not self.files:
            messagebox.showwarning("No file", "Please select PDF file first.")
            return

        self.btn_process.configure(state="disabled")
        self.set_status("Processing...")
        self.log("▶ Starting analysis...")

        def worker():
            try:
                all_dfs = []
                metas = []
                for path in self.files:
                    self.after(0, lambda p=path: self.log(f"\nProcessing: {os.path.basename(p)}"))
                    df_one, meta_one = parse_cr_pdf(path, return_meta=True)
                    all_dfs.append(df_one)
                    metas.append(meta_one)

                self.df = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()
                self.meta = metas[0] if metas else {}

                if not self.df.empty and "enrollment_no" in self.df.columns:
                    self.df["__name_len"] = self.df["name"].astype(str).str.len()
                    self.df = self.df.sort_values(["enrollment_no", "__name_len"], ascending=[True, False])
                    self.df = self.df.drop_duplicates(subset=["enrollment_no"], keep="first").drop(columns=["__name_len"], errors="ignore")
                    self.df = self.df.reset_index(drop=True)

                self.analysis = self._run_analysis(self.df)
                self.subject_options = ["All Subjects"] + [x["display_name"] for x in self._subject_info_list_export(self._subject_groups())]
                self.after(0, lambda: self.graph_subject_menu.configure(values=self.subject_options))
                self.after(0, lambda: self.graph_subject_var.set("All Subjects"))

                self.after(0, lambda: self.lbl_students.configure(text=str(self.analysis.get('students', 0))))
                self.after(0, lambda: self.lbl_pass.configure(text=str(self.analysis.get('pass', 0))))
                self.after(0, lambda: self.lbl_fail.configure(text=str(self.analysis.get('fail', 0))))
                avgp = self.analysis.get("avg_percent", None)
                self.after(0, lambda: self.lbl_avg.configure(text=f"{avgp:.2f}" if avgp is not None else "--"))
                self.after(0, lambda: self.log("\n✅ Analysis Completed!"))
                self.after(0, lambda: self.log(f"Top 10 ready: {len(self.analysis.get('top10', []))} students"))
                self.after(0, lambda: self.btn_graph.configure(state="normal"))
                self.after(0, lambda: self.btn_top10.configure(state="normal"))
                self.after(0, lambda: self.btn_students.configure(state="normal"))
                self.after(0, lambda: self.btn_export_excel.configure(state="normal"))
                self.after(0, lambda: self.graph_hint.configure(text="Select All Subjects or one subject to view graph."))
                self.after(0, lambda: self.set_status("Done"))
                self.after(0, lambda: self.show_page("home"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
                self.after(0, lambda: self.set_status("Error"))
            finally:
                self.after(0, lambda: self.btn_process.configure(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _run_analysis(self, df):
        if df is None or df.empty:
            return {"students": 0, "pass": 0, "fail": 0, "avg_percent": None, "top10": []}
        d = df.copy()
        d["result"] = d["result"].fillna("UNKNOWN").astype(str).str.upper()
        avgp = pd.to_numeric(d["percentage"], errors="coerce").mean()
        d["__total"] = pd.to_numeric(d["total_marks_550"], errors="coerce").fillna(0)
        d["__pct"] = pd.to_numeric(d["percentage"], errors="coerce").fillna(0)
        top = d[d["result"] == "PASS"].sort_values(["__total", "__pct"], ascending=[False, False]).head(10).reset_index(drop=True)
        top10 = top.drop(columns=["__total", "__pct"], errors="ignore").to_dict(orient="records")
        return {
            "students": len(d),
            "pass": int((d["result"] == "PASS").sum()),
            "fail": int((d["result"] == "FAIL").sum()),
            "avg_percent": float(avgp) if pd.notna(avgp) else None,
            "top10": top10
        }

    # ---------- Export Helpers ----------
    def _subject_groups(self):
        groups = []
        for s in (self.meta or {}).get("subjects", []):
            pref = str(s.get("col_prefix", "")).strip()
            groups.append((
                str(s.get("name", "")).strip(),
                str(s.get("cs_num", "")).strip(),
                f"{pref}_the_total",
                f"{pref}_int_total",
                f"{pref}_pra_total",
                f"{pref}_total",
            ))
        return groups

    def _subject_info_list_export(self, groups):
        out = []
        for meta_name, cs_num, col_the, col_int, col_pra, col_tot in groups:
            pref = col_tot[:-6] if col_tot.endswith("_total") else ""
            display_name = meta_name or (cs_num if cs_num else pref.upper())
            out.append({
                "display_name": display_name,
                "the_col": col_the,
                "int_col": col_int,
                "pra_col": col_pra,
                "tot_col": col_tot,
                "grade_col": f"{pref}_grade_token",
                "status_col": f"{col_tot}_status",
            })
        return out

    def _subject_status_value_export(self, row, info):
        def norm(v):
            return str(v).strip().upper()

        the_v = norm(row.get(info["the_col"], ""))
        int_v = norm(row.get(info["int_col"], ""))
        pra_v = norm(row.get(info["pra_col"], "")) if info["pra_col"] else ""
        grade_v = norm(row.get(info["grade_col"], ""))

        if the_v in ("AB", "A B", "AL", "ABSENT") or int_v in ("AB", "A B", "AL", "ABSENT") or pra_v in ("AB", "A B", "AL", "ABSENT") or grade_v == "AB":
            return "ABSENT"

        if grade_v:
            parts = grade_v.split("/")
            if len(parts) >= 2:
                return "FAIL" if parts[1].strip().upper() == "F" else "PASS"

        total_v = pd.to_numeric(pd.Series([row.get(info["tot_col"], "")]), errors="coerce").iloc[0]
        if pd.notna(total_v):
            return "PASS"
        return ""

    def _subject_status_dataframe_export(self, df_raw, groups):
        if df_raw is None or df_raw.empty:
            return pd.DataFrame()
        df2 = df_raw.copy()
        for info in self._subject_info_list_export(groups):
            df2[info["status_col"]] = df2.apply(lambda r: self._subject_status_value_export(r, info), axis=1)
        return df2

    def _build_students_sheet(self, df_raw, groups):
        df2 = self._subject_status_dataframe_export(df_raw, groups)
        out = pd.DataFrame()
        out["Seat No."] = df2.get("enrollment_no", "")
        out["Name"] = df2.get("name", "")

        for info in self._subject_info_list_export(groups):
            disp = info["display_name"]
            out[f"{disp} (THE TOTAL)"] = df2.get(info["the_col"], "")
            out[f"{disp} (INT TOTAL)"] = df2.get(info["int_col"], "")
            if info["pra_col"] in df2.columns:
                pra_series = df2.get(info["pra_col"], "")
                if pd.Series(pra_series).astype(str).str.strip().ne("").any():
                    out[f"{disp} (PRA TOTAL)"] = pra_series
            out[f"{disp} (Total Marks)"] = df2.get(info["tot_col"], "")
            out[f"{disp} (Grade Token)"] = df2.get(info["grade_col"], "")
            out[f"{disp} (Status)"] = df2.get(info["status_col"], "")

        out["Pass/Fail"] = df2.get("result", "")
        out["Total"] = df2.get("total_marks_550", "")
        out["Percentage"] = df2.get("percentage", "")
        out["SGPA"] = df2.get("sgpa", "")
        out["CGPA"] = df2.get("cgpa", "")
        return out.fillna("")

    def _build_top10_sheet(self, df_raw, groups):
        top = df_raw.copy()
        top["result"] = top["result"].fillna("").astype(str).str.upper()
        top = top[top["result"] == "PASS"].copy()
        top["__total"] = pd.to_numeric(top["total_marks_550"], errors="coerce").fillna(0)
        top["__pct"] = pd.to_numeric(top["percentage"], errors="coerce").fillna(0)
        top = top.sort_values(["__total", "__pct"], ascending=[False, False]).head(10).reset_index(drop=True)

        out = pd.DataFrame()
        out["Rank"] = range(1, len(top) + 1)
        out["Seat No."] = top.get("enrollment_no", "")
        out["Name"] = top.get("name", "")
        for info in self._subject_info_list_export(groups):
            out[f"{info['display_name']} (Total Marks)"] = top.get(info["tot_col"], "")
        out["Total"] = top.get("total_marks_550", "")
        out["Percentage"] = top.get("percentage", "")
        out["SGPA"] = top.get("sgpa", "")
        out["CGPA"] = top.get("cgpa", "")
        return out.fillna("")

    def _build_subjectwise_sheet(self, df_raw, groups):
        df2 = self._subject_status_dataframe_export(df_raw, groups)
        rows = []
        for info in self._subject_info_list_export(groups):
            s = df2.get(info["status_col"], pd.Series(dtype=str)).astype(str).str.upper()
            passed_marks = pd.to_numeric(df2.loc[s == "PASS", info["tot_col"]], errors="coerce")

            pass_avg = ""
            if not passed_marks.empty:
                avg_marks = float(passed_marks.mean())
                all_marks = pd.to_numeric(df2[info["tot_col"]], errors="coerce")
                max_seen = all_marks.max()
                subject_max = None
                if pd.notna(max_seen):
                    if max_seen <= 50:
                        subject_max = 50.0
                    elif max_seen <= 70:
                        subject_max = 70.0
                    else:
                        subject_max = 100.0
                pass_avg = round((avg_marks / subject_max) * 100, 2) if subject_max else round(avg_marks, 2)

            rows.append({
                "Subject": info["display_name"],
                "Pass": int((s == "PASS").sum()),
                "Fail": int((s == "FAIL").sum()),
                "Absent": int((s == "ABSENT").sum()),
                "Total Students": len(df2),
                "Passing Student Avg. %": pass_avg
            })
        return pd.DataFrame(rows)

    def export_excel(self):
        if self.df.empty:
            messagebox.showwarning("Export", "No data to export.")
            return

        out = filedialog.asksaveasfilename(
            title="Save Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")]
        )
        if not out:
            return

        try:
            groups = self._subject_groups()
            college = get_college_name()

            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                self._build_students_sheet(self.df, groups).to_excel(writer, index=False, sheet_name="Students", startrow=2)
                self._build_top10_sheet(self.df, groups).to_excel(writer, index=False, sheet_name="Top10", startrow=2)
                self._build_subjectwise_sheet(self.df, groups).to_excel(writer, index=False, sheet_name="SubjectWise", startrow=2)

                # Add college name header to each sheet
                if college:
                    from openpyxl.styles import Font, Alignment, PatternFill
                    for sheet_name in ["Students", "Top10", "SubjectWise"]:
                        ws = writer.sheets[sheet_name]

                        # Row 1 — College name (big, bold, blue background)
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ws.max_column, 6))
                        cell = ws.cell(row=1, column=1)
                        cell.value = college.upper()
                        cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                        ws.row_dimensions[1].height = 28

                        # Row 2 — Sheet title (e.g. "Students Report")
                        title_map = {
                            "Students":    "Students Result Report",
                            "Top10":       "Top 10 Students Report",
                            "SubjectWise": "Subject Wise Analysis Report",
                        }
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ws.max_column, 6))
                        cell2 = ws.cell(row=2, column=1)
                        cell2.value = title_map.get(sheet_name, sheet_name)
                        cell2.font = Font(name="Calibri", size=12, bold=True, color="1E3A5F")
                        cell2.alignment = Alignment(horizontal="center", vertical="center")
                        cell2.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                        ws.row_dimensions[2].height = 20

            try:
                os.startfile(out)
            except Exception:
                pass
            messagebox.showinfo("Export", f"Excel exported successfully:\n{out}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ---------- Utils ----------
    def log(self, msg: str):
        self.logbox.insert("end", msg + "\n")
        self.logbox.see("end")

    def set_status(self, msg: str):
        self.status.configure(text=f"Status: {msg}")

    def clear_all(self):
        self.files = []
        self.df = pd.DataFrame()
        self.meta = {}
        self.analysis = {}
        self.subject_options = []
        self.selected_subjects = []

        self.card_students.set_value("--")
        self.card_pass.set_value("--")
        self.card_fail.set_value("--")
        self.card_avg.set_value("--")

        self.btn_process.configure(state="disabled")
        self.btn_graph.configure(state="disabled")
        self.btn_top10.configure(state="disabled")
        self.btn_students.configure(state="disabled")
        self.btn_export_excel.configure(state="disabled")

        self.graph_subject_menu.configure(values=["All Subjects"])
        self.graph_subject_var.set("All Subjects")
        self.graph_hint.configure(text="Process PDFs first, then choose a subject.")

        self.set_status("Idle")
        self.logbox.delete("1.0", "end")
        self.log("Cleared. Ready.")
        self.show_page("home")


if __name__ == "__main__":
    app = ResultAnalyzerApp()
    app.mainloop()
